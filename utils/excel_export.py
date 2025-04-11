from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def export_protocol_to_excel(file_path, data, warunki_przed, warunki_po, wyposazenie):
    wb = Workbook()
    ws = wb.active
    ws.title = "Protokół"


    ws.merge_cells("A1:E1")
    ws["A1"] = "PROTOKÓŁ POMIAROWY"
    ws["A1"].style = "Title"


    ws["A3"] = "Warunki środowiskowe (PRZED):"
    ws["B3"] = warunki_przed
    ws["A4"] = "Warunki środowiskowe (PO):"
    ws["B4"] = warunki_po


    ws["A6"] = "Użyte wyposażenie:"
    ws["B6"] = wyposazenie


    headers = ["Nazwa", "Wartość", "Jednostka", "Minimalna", "Maksymalna"]
    ws.append([""] * 5)
    ws.append(headers)

    for row in data:
        ws.append(row)


    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2


    filename = file_path or f"protokol_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename
